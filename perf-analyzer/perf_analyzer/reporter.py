"""리포트 생성기 — Excel + 마크다운.

generate-summary.py의 create_excel() 로직을 재사용하되,
병목 분석 결과와 런북을 추가로 포함한다.
"""

import json
from datetime import datetime
from pathlib import Path
from typing import Optional

from .models import AnalysisReport, Bottleneck, K6ScenarioResult


# ──────────────────────────────────────────────────────────
# 마크다운 리포트
# ──────────────────────────────────────────────────────────

def generate_markdown_report(report: AnalysisReport, output_dir: str) -> Path:
    """분석 결과를 마크다운 런북으로 출력."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    date_str = report.test_date or datetime.now().strftime("%Y-%m-%d")
    filename = f"analysis-report_{date_str}.md"
    filepath = output_path / filename

    lines = []
    lines.append(f"# 부하테스트 분석 리포트")
    lines.append(f"")
    lines.append(f"> 생성: {report.analysis_time}")
    lines.append(f"> 테스트 날짜: {date_str}")
    if report.prometheus_url:
        lines.append(f"> Prometheus: {report.prometheus_url}")
    lines.append(f"")

    # ── 시나리오 요약 ──
    lines.append(f"## 시나리오 요약")
    lines.append(f"")
    lines.append(f"| 시나리오 | 판정 | 요청 수 | P50 (ms) | P95 (ms) | P99 (ms) | 에러율 | TPS | Peak TPS | 핵심 지표 |")
    lines.append(f"|----------|------|---------|----------|----------|----------|--------|-----|----------|----------|")

    for s in report.scenarios:
        verdict = "✅ PASS" if s.passed else "❌ FAIL"
        lines.append(
            f"| {s.scenario} | {verdict} | {s.total_requests:,} | "
            f"{s.p50_ms:.1f} | {s.p95_ms:.1f} | {s.p99_ms:.1f} | {s.error_rate} | "
            f"{s.tps:.1f} | {s.peak_tps:.1f} | {s.key_metric} |"
        )
    lines.append(f"")

    # ── HTTP 상태 코드 분포 ──
    has_status_dist = any(s.http_status_dist for s in report.scenarios)
    if has_status_dist:
        lines.append(f"### HTTP 상태 코드 분포")
        lines.append(f"")
        for s in report.scenarios:
            if s.http_status_dist:
                dist_str = ", ".join(f"{k}: {v}" for k, v in s.http_status_dist.items())
                lines.append(f"- **{s.scenario}**: {dist_str}")
        lines.append(f"")

    # ── 시간대별 TPS ──
    has_tps_timeline = any(s.tps_timeline for s in report.scenarios)
    if has_tps_timeline:
        lines.append(f"### 시간대별 TPS (추정)")
        lines.append(f"")
        for s in report.scenarios:
            if s.tps_timeline:
                lines.append(f"**{s.scenario}** (총 {s.duration_seconds:.0f}초):")
                for entry in s.tps_timeline:
                    if len(entry) >= 3:
                        lines.append(f"- {entry[2]}: ~{entry[1]} req/s")
                lines.append(f"")
        lines.append(f"")

    # ── 병목 분석 결과 ──
    if report.bottlenecks:
        lines.append(f"## 식별된 병목 ({len(report.bottlenecks)}건)")
        lines.append(f"")

        for i, b in enumerate(report.bottlenecks, 1):
            severity_emoji = {
                "critical": "🔴",
                "high": "🟠",
                "medium": "🟡",
                "low": "🟢",
            }.get(b.severity, "⚪")

            lines.append(f"### {i}. {severity_emoji} [{b.severity.upper()}] {b.category}")
            lines.append(f"")
            lines.append(f"**설명**: {b.description}")
            lines.append(f"")

            if b.evidence:
                lines.append(f"**근거**:")
                for e in b.evidence:
                    lines.append(f"- {e}")
                lines.append(f"")

            if b.incident_id:
                lines.append(f"**관련 런북**: {b.incident_id}")
                lines.append(f"")

            if b.runbook_summary:
                lines.append(f"**요약**: {b.runbook_summary}")
                lines.append(f"")

            if b.recommended_actions:
                lines.append(f"**권장 조치**:")
                for a in b.recommended_actions:
                    lines.append(f"1. {a}")
                lines.append(f"")

            if b.metric_name:
                lines.append(f"**메트릭**: `{b.metric_name}` — 임계치: {b.threshold}, 실측: {b.actual_value:.3f}")
                lines.append(f"")

            lines.append(f"---")
            lines.append(f"")

    else:
        lines.append(f"## 병목 분석 결과")
        lines.append(f"")
        lines.append(f"식별된 병목이 없습니다. 모든 지표가 정상 범위 내입니다.")
        lines.append(f"")

    # ── k6 진단 (원본) ──
    diag_scenarios = [s for s in report.scenarios if s.diagnostics]
    if diag_scenarios:
        lines.append(f"## k6 자체 진단")
        lines.append(f"")
        for s in diag_scenarios:
            lines.append(f"### {s.scenario}")
            for d in s.diagnostics:
                lines.append(f"")
                lines.append(f"**증상**: {d.get('symptom', '')}")
                for c in d.get("causes", []):
                    lines.append(f"- 원인: {c.get('cause', '')}")
                    lines.append(f"  - 확인: {c.get('check', '')}")
            lines.append(f"")

    # ── 다음 단계 ──
    lines.append(f"## 다음 단계")
    lines.append(f"")
    if report.bottlenecks:
        lines.append(f"1. 위 권장 조치를 severity 순서대로 적용")
        lines.append(f"2. 조치 후 동일 시나리오 재실행")
        lines.append(f"3. `perf-analyzer compare`로 개선 전/후 수치 비교")
        lines.append(f"4. 결과를 Notion/Slack에 공유")
    else:
        lines.append(f"1. 추가 시나리오(좌석 경쟁, E2E 흐름) k6 스크립트 작성")
        lines.append(f"2. VU 수를 늘려 한계점 탐색")
        lines.append(f"3. SLI/SLO 기준선 업데이트")
    lines.append(f"")
    lines.append(f"---")
    lines.append(f"*Generated by perf-analyzer v0.1.0*")

    filepath.write_text("\n".join(lines), encoding="utf-8")
    return filepath


# ──────────────────────────────────────────────────────────
# Excel 리포트
# ──────────────────────────────────────────────────────────

def generate_excel_report(report: AnalysisReport, output_dir: str) -> Path:
    """분석 결과를 Excel로 출력.

    generate-summary.py의 create_excel() 스타일을 재사용하되,
    '병목 분석' 시트를 추가한다.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        import subprocess, sys
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
        )
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    date_str = report.test_date or datetime.now().strftime("%Y-%m-%d")
    filename = f"analysis-report_{date_str}.xlsx"
    filepath = output_path / filename

    wb = openpyxl.Workbook()

    # 스타일 (generate-summary.py에서 가져옴)
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    pass_font = Font(bold=True, color="16A34A")
    fail_font = Font(bold=True, color="DC2626")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="6B7280")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    # ── Summary 시트 ──
    ws = wb.active
    ws.title = "시나리오 요약"
    ws.cell(row=1, column=1, value="부하테스트 분석 리포트").font = title_font
    ws.cell(row=2, column=1, value=f"테스트 날짜: {date_str}").font = subtitle_font
    ws.cell(row=3, column=1, value=f"분석 시각: {report.analysis_time}").font = subtitle_font

    headers = ["시나리오", "판정", "요청 수", "P50 (ms)", "P95 (ms)", "P99 (ms)",
               "Max (ms)", "에러율", "TPS", "Peak TPS", "테스트 시간(s)", "핵심 지표"]
    hr = 5
    for i, h in enumerate(headers, 1):
        ws.cell(row=hr, column=i, value=h)
    style_header(ws, hr, len(headers))

    for ri, s in enumerate(report.scenarios, hr + 1):
        ws.cell(row=ri, column=1, value=s.scenario).border = thin_border
        verdict_cell = ws.cell(row=ri, column=2, value="PASS" if s.passed else "FAIL")
        verdict_cell.border = thin_border
        verdict_cell.font = pass_font if s.passed else fail_font
        verdict_cell.fill = pass_fill if s.passed else fail_fill
        ws.cell(row=ri, column=3, value=s.total_requests).border = thin_border
        ws.cell(row=ri, column=4, value=round(s.p50_ms, 1)).border = thin_border
        ws.cell(row=ri, column=5, value=round(s.p95_ms, 1)).border = thin_border
        ws.cell(row=ri, column=6, value=round(s.p99_ms, 1)).border = thin_border
        ws.cell(row=ri, column=7, value=round(s.max_ms, 1)).border = thin_border
        ws.cell(row=ri, column=8, value=s.error_rate).border = thin_border
        ws.cell(row=ri, column=9, value=round(s.tps, 1)).border = thin_border
        ws.cell(row=ri, column=10, value=round(s.peak_tps, 1)).border = thin_border
        ws.cell(row=ri, column=11, value=round(s.duration_seconds, 0)).border = thin_border
        ws.cell(row=ri, column=12, value=s.key_metric).border = thin_border

    for i, w in enumerate([30, 8, 12, 12, 12, 12, 12, 18, 10, 12, 12, 25], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── HTTP 상태 코드 분포 시트 ──
    has_dist = any(s.http_status_dist for s in report.scenarios)
    if has_dist:
        ws_dist = wb.create_sheet(title="HTTP 상태 분포")
        ws_dist.cell(row=1, column=1, value="HTTP 상태 코드 분포").font = title_font

        # 모든 상태 코드 키 수집
        all_codes = set()
        for s in report.scenarios:
            all_codes.update(s.http_status_dist.keys())
        sorted_codes = sorted(all_codes)

        dist_headers = ["시나리오"] + sorted_codes + ["합계"]
        r_dist = 3
        for i, h in enumerate(dist_headers, 1):
            ws_dist.cell(row=r_dist, column=i, value=h)
        style_header(ws_dist, r_dist, len(dist_headers))

        for s in report.scenarios:
            r_dist += 1
            ws_dist.cell(row=r_dist, column=1, value=s.scenario).border = thin_border
            total_in_dist = 0
            for ci, code in enumerate(sorted_codes, 2):
                val = s.http_status_dist.get(code, 0)
                total_in_dist += val
                ws_dist.cell(row=r_dist, column=ci, value=val).border = thin_border
            ws_dist.cell(row=r_dist, column=len(sorted_codes) + 2, value=total_in_dist).border = thin_border

    # ── 병목 분석 시트 ──
    if report.bottlenecks:
        ws_bn = wb.create_sheet(title="병목 분석")
        ws_bn.cell(row=1, column=1, value="식별된 병목").font = title_font
        ws_bn.cell(row=2, column=1, value=f"총 {len(report.bottlenecks)}건").font = subtitle_font

        bn_headers = ["심각도", "카테고리", "서비스", "설명", "관련 런북", "권장 조치"]
        hr2 = 4
        for i, h in enumerate(bn_headers, 1):
            ws_bn.cell(row=hr2, column=i, value=h)
        style_header(ws_bn, hr2, len(bn_headers))

        severity_fills = {
            "critical": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
            "high": PatternFill(start_color="FED7AA", end_color="FED7AA", fill_type="solid"),
            "medium": PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid"),
            "low": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        }

        for ri, b in enumerate(report.bottlenecks, hr2 + 1):
            sev_cell = ws_bn.cell(row=ri, column=1, value=b.severity.upper())
            sev_cell.border = thin_border
            sev_cell.fill = severity_fills.get(b.severity, PatternFill())
            sev_cell.font = Font(bold=True)
            ws_bn.cell(row=ri, column=2, value=b.category).border = thin_border
            # description에서 서비스 이름 추출
            service = b.description.split(":")[0] if ":" in b.description else ""
            ws_bn.cell(row=ri, column=3, value=service).border = thin_border
            ws_bn.cell(row=ri, column=4, value=b.description).border = thin_border
            ws_bn.cell(row=ri, column=5, value=b.incident_id).border = thin_border
            actions = " → ".join(b.recommended_actions[:3]) if b.recommended_actions else ""
            ws_bn.cell(row=ri, column=6, value=actions).border = thin_border

        for i, w in enumerate([12, 18, 20, 60, 12, 60], 1):
            ws_bn.column_dimensions[get_column_letter(i)].width = w

    # ── 시나리오별 상세 시트 (generate-summary.py 스타일) ──
    for s in report.scenarios:
        raw = s.raw
        if not raw:
            continue
        sheet_name = s.scenario[:31]
        ws_d = wb.create_sheet(title=sheet_name)
        ws_d.cell(row=1, column=1, value=f"{s.scenario} 상세").font = title_font
        ws_d.cell(row=2, column=1, value=f"실행: {s.timestamp}").font = subtitle_font

        r = 4
        # 결과 지표
        results_data = raw.get("results", {})
        if results_data:
            ws_d.cell(row=r, column=1, value="결과 지표").font = Font(bold=True, size=11)
            r += 1
            for i, h in enumerate(["지표", "값"], 1):
                ws_d.cell(row=r, column=i, value=h)
            style_header(ws_d, r, 2)
            r += 1
            for key, val in results_data.items():
                ws_d.cell(row=r, column=1, value=key).border = thin_border
                ws_d.cell(row=r, column=2, value=val).border = thin_border
                r += 1

        # 레이턴시
        lat = raw.get("latency", {})
        if lat:
            r += 1
            ws_d.cell(row=r, column=1, value="레이턴시 (ms)").font = Font(bold=True, size=11)
            r += 1
            for i, h in enumerate(["구분", "P50", "P95", "P99"], 1):
                ws_d.cell(row=r, column=i, value=h)
            style_header(ws_d, r, 4)
            r += 1
            for lname, ldata in lat.items():
                ws_d.cell(row=r, column=1, value=lname).border = thin_border
                ws_d.cell(row=r, column=2, value=ldata.get("p50", 0)).border = thin_border
                ws_d.cell(row=r, column=3, value=ldata.get("p95", 0)).border = thin_border
                ws_d.cell(row=r, column=4, value=ldata.get("p99", 0)).border = thin_border
                r += 1

        ws_d.column_dimensions["A"].width = 35
        ws_d.column_dimensions["B"].width = 25

    wb.save(str(filepath))
    return filepath
