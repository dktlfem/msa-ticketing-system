#!/usr/bin/env python3
"""
k6 부하테스트 결과 Excel 요약 생성기

사용법:
  python3 generate-summary.py results/

결과 디렉토리 내 *.json 파일을 읽어 Excel(.xlsx) 요약을 생성한다.
같은 시나리오의 JSON이 여러 개 있으면 최신(timestamp 기준)만 사용한다.
- Summary 시트: 전체 시나리오 p95/p99, 에러율, TPS 한눈 요약
- 시나리오별 상세 시트
"""

import json
import os
import sys
import subprocess
from datetime import datetime
from pathlib import Path

def ensure_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("  openpyxl 설치 중...")
        subprocess.check_call(
            [sys.executable, "-m", "pip", "install", "openpyxl", "-q"],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
        )
        import openpyxl
        return openpyxl


def load_json_results(result_dir):
    """json/ 서브디렉토리의 *.json 파일을 로드, 없으면 루트에서 탐색 (하위 호환)"""
    all_results = []
    json_dir = Path(result_dir) / "json"
    search_dir = json_dir if json_dir.exists() else Path(result_dir)
    for f in sorted(search_dir.glob("*.json")):
        try:
            with open(f, "r", encoding="utf-8") as fh:
                data = json.load(fh)
            if "scenario" in data:
                data["_file"] = f.name
                all_results.append(data)
        except (json.JSONDecodeError, KeyError):
            continue

    # 같은 시나리오가 여러 개면 timestamp가 가장 최신인 것만 사용
    latest = {}
    for r in all_results:
        scenario = r.get("scenario", "unknown")
        ts = r.get("timestamp", "")
        if scenario not in latest or ts > latest[scenario].get("timestamp", ""):
            latest[scenario] = r

    return list(latest.values())


def extract_summary_row(report):
    """JSON 리포트에서 요약 행 추출"""
    scenario = report.get("scenario", "unknown")
    timestamp = report.get("timestamp", "")
    results = report.get("results", {})
    latency = report.get("latency", {})
    pass_fail = "PASS" if report.get("pass") else "FAIL"

    # 시나리오별 핵심 지표 추출
    row = {
        "scenario": scenario,
        "timestamp": timestamp,
        "pass": pass_fail,
        "total_requests": results.get("totalRequests", 0),
    }

    if scenario == "scenario1-rate-limiter":
        row["p95_ms"] = latency.get("allowed", {}).get("p95", 0)
        row["p99_ms"] = latency.get("allowed", {}).get("p99", 0)
        row["error_rate"] = f"{results.get('rateLimitedPercent', 0)}% (429)"
        row["tps"] = results.get("avgAllowedTps", 0)
        row["key_metric"] = f"차단율 {results.get('rateLimitedPercent', 0)}%"

    elif scenario == "scenario2-circuit-breaker":
        row["p95_ms"] = latency.get("closed", {}).get("p95", 0)
        row["p99_ms"] = latency.get("closed", {}).get("p99", 0)
        fb_pct = results.get("fallbackPercent", 0)
        row["error_rate"] = f"{fb_pct}% (fallback)"
        total = results.get("totalRequests", 1)
        duration = 105  # baseline 15 + fault 60 + recovery 30
        row["tps"] = round(total / duration, 2) if total else 0
        row["key_metric"] = f"FB {results.get('fallbackTotal', 0)}건"
        # 순수 fallback p95
        pure_fb_p95 = latency.get("pureFallback", {}).get("p95", 0)
        row["pure_fb_p95_ms"] = pure_fb_p95

    elif scenario == "scenario3-bulkhead":
        row["p95_ms"] = latency.get("passed", {}).get("p95", 0)
        row["p99_ms"] = latency.get("passed", {}).get("p99", 0)
        rej_pct = results.get("bulkheadRejectedPercent", 0)
        row["error_rate"] = f"{rej_pct}% (rejected)"
        total = results.get("totalRequests", 1)
        duration = 80  # baseline 15 + saturation 35 + recovery 30
        row["tps"] = round(total / duration, 2) if total else 0
        row["key_metric"] = f"BH거절 {results.get('bulkheadRejected', 0)}건"
        rej_p95 = latency.get("bulkheadRejected", {}).get("p95", 0)
        row["rejected_p95_ms"] = rej_p95

    else:
        row["p95_ms"] = 0
        row["p99_ms"] = 0
        row["error_rate"] = "-"
        row["tps"] = 0
        row["key_metric"] = "-"

    return row


def create_excel(results, result_dir):
    openpyxl = ensure_openpyxl()
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today = datetime.now().strftime("%Y-%m-%d")
    wb = openpyxl.Workbook()

    # ── 스타일 정의 ─────────────────────────────────────────
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    pass_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    pass_font = Font(bold=True, color="16A34A")
    fail_font = Font(bold=True, color="DC2626")
    title_font = Font(bold=True, size=14)
    subtitle_font = Font(bold=True, size=11, color="6B7280")
    num_font = Font(name="Consolas", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    def style_header(ws, row, cols):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

    def style_data_cell(ws, row, col, is_num=False):
        cell = ws.cell(row=row, column=col)
        cell.border = thin_border
        if is_num:
            cell.font = num_font
            cell.alignment = Alignment(horizontal="right")
        return cell

    # ── Summary 시트 ────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"

    ws.cell(row=1, column=1, value="SCG Hardening 부하테스트 요약").font = title_font
    ws.cell(row=2, column=1, value=f"테스트 날짜: {today}").font = subtitle_font
    ws.cell(row=3, column=1, value=f"생성 시각: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}").font = subtitle_font

    # 요약 테이블
    summary_headers = [
        "시나리오", "판정", "전체 요청", "P95 (ms)", "P99 (ms)",
        "에러율/차단율", "TPS (req/s)", "핵심 지표",
    ]
    header_row = 5
    for i, h in enumerate(summary_headers, 1):
        ws.cell(row=header_row, column=i, value=h)
    style_header(ws, header_row, len(summary_headers))

    rows = [extract_summary_row(r) for r in results]

    for ri, row in enumerate(rows, header_row + 1):
        ws.cell(row=ri, column=1, value=row["scenario"]).border = thin_border

        pass_cell = ws.cell(row=ri, column=2, value=row["pass"])
        pass_cell.border = thin_border
        pass_cell.alignment = Alignment(horizontal="center")
        if row["pass"] == "PASS":
            pass_cell.font = pass_font
            pass_cell.fill = pass_fill
        else:
            pass_cell.font = fail_font
            pass_cell.fill = fail_fill

        style_data_cell(ws, ri, 3, True).value = row["total_requests"]
        style_data_cell(ws, ri, 4, True).value = row["p95_ms"]
        style_data_cell(ws, ri, 5, True).value = row["p99_ms"]
        style_data_cell(ws, ri, 6).value = row["error_rate"]
        style_data_cell(ws, ri, 7, True).value = row["tps"]
        style_data_cell(ws, ri, 8).value = row["key_metric"]

    # 열 너비 조정
    col_widths = [30, 8, 12, 12, 12, 20, 12, 25]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── 시나리오별 상세 시트 ────────────────────────────────
    for report in results:
        scenario = report.get("scenario", "unknown")
        # 시트 이름은 31자 제한
        sheet_name = scenario[:31]
        ws_detail = wb.create_sheet(title=sheet_name)

        ws_detail.cell(row=1, column=1, value=f"{scenario} 상세").font = title_font
        ws_detail.cell(row=2, column=1, value=f"실행: {report.get('timestamp', '')}").font = subtitle_font
        pass_text = "PASS" if report.get("pass") else "FAIL"
        pass_cell = ws_detail.cell(row=2, column=4, value=pass_text)
        if report.get("pass"):
            pass_cell.font = pass_font
        else:
            pass_cell.font = fail_font

        # 설정 정보
        config = report.get("config", {})
        r = 4
        ws_detail.cell(row=r, column=1, value="설정").font = Font(bold=True, size=11)
        r += 1
        detail_headers = ["항목", "값"]
        for i, h in enumerate(detail_headers, 1):
            ws_detail.cell(row=r, column=i, value=h)
        style_header(ws_detail, r, 2)
        r += 1

        ws_detail.cell(row=r, column=1, value="SCG URL").border = thin_border
        ws_detail.cell(row=r, column=2, value=config.get("scgBaseUrl", "")).border = thin_border
        r += 1
        ws_detail.cell(row=r, column=1, value="Target Path").border = thin_border
        ws_detail.cell(row=r, column=2, value=config.get("targetPath", "")).border = thin_border
        r += 1

        # Phase 정보
        phases = report.get("phases", {})
        if phases:
            r += 1
            ws_detail.cell(row=r, column=1, value="Phase 구성").font = Font(bold=True, size=11)
            r += 1
            phase_headers = ["Phase", "VUs", "Duration", "Sleep", "Purpose"]
            for i, h in enumerate(phase_headers, 1):
                ws_detail.cell(row=r, column=i, value=h)
            style_header(ws_detail, r, len(phase_headers))
            r += 1
            for pname, pdata in phases.items():
                ws_detail.cell(row=r, column=1, value=pname).border = thin_border
                ws_detail.cell(row=r, column=2, value=pdata.get("vus", "")).border = thin_border
                ws_detail.cell(row=r, column=3, value=pdata.get("duration", "")).border = thin_border
                sleep_val = pdata.get("sleepMs", pdata.get("sleep", ""))
                ws_detail.cell(row=r, column=4, value=f"{sleep_val}ms" if isinstance(sleep_val, (int, float)) else str(sleep_val)).border = thin_border
                ws_detail.cell(row=r, column=5, value=pdata.get("purpose", "")).border = thin_border
                r += 1

        # 결과 지표
        results_data = report.get("results", {})
        if results_data:
            r += 1
            ws_detail.cell(row=r, column=1, value="결과 지표").font = Font(bold=True, size=11)
            r += 1
            metric_headers = ["지표", "값"]
            for i, h in enumerate(metric_headers, 1):
                ws_detail.cell(row=r, column=i, value=h)
            style_header(ws_detail, r, 2)
            r += 1
            for key, val in results_data.items():
                ws_detail.cell(row=r, column=1, value=key).border = thin_border
                # dict/list 등 복합 타입은 JSON 문자열로 변환
                if isinstance(val, (dict, list)):
                    cell_val = json.dumps(val, ensure_ascii=False)
                else:
                    cell_val = val
                style_data_cell(ws_detail, r, 2, True).value = cell_val
                r += 1

        # 레이턴시
        latency = report.get("latency", {})
        if latency:
            r += 1
            ws_detail.cell(row=r, column=1, value="레이턴시 (ms)").font = Font(bold=True, size=11)
            r += 1
            lat_headers = ["구분", "P50", "P95", "P99"]
            for i, h in enumerate(lat_headers, 1):
                ws_detail.cell(row=r, column=i, value=h)
            style_header(ws_detail, r, len(lat_headers))
            r += 1
            for lname, ldata in latency.items():
                ws_detail.cell(row=r, column=1, value=lname).border = thin_border
                style_data_cell(ws_detail, r, 2, True).value = ldata.get("p50", 0)
                style_data_cell(ws_detail, r, 3, True).value = ldata.get("p95", 0)
                style_data_cell(ws_detail, r, 4, True).value = ldata.get("p99", 0)
                r += 1

        # Threshold 판정
        thresholds = report.get("thresholds", {})
        if thresholds:
            r += 1
            ws_detail.cell(row=r, column=1, value="Threshold 판정").font = Font(bold=True, size=11)
            r += 1
            th_headers = ["Threshold", "Expression", "결과"]
            for i, h in enumerate(th_headers, 1):
                ws_detail.cell(row=r, column=i, value=h)
            style_header(ws_detail, r, len(th_headers))
            r += 1
            for tname, tdata in thresholds.items():
                for expr, result in tdata.items():
                    ws_detail.cell(row=r, column=1, value=tname).border = thin_border
                    ws_detail.cell(row=r, column=2, value=expr).border = thin_border
                    ok = result.get("ok", result) if isinstance(result, dict) else result
                    result_cell = ws_detail.cell(row=r, column=3, value="PASS" if ok else "FAIL")
                    result_cell.border = thin_border
                    result_cell.font = pass_font if ok else fail_font
                    result_cell.fill = pass_fill if ok else fail_fill
                    r += 1

        # 진단
        diagnostics = report.get("diagnostics", [])
        if diagnostics:
            r += 1
            ws_detail.cell(row=r, column=1, value="진단").font = Font(bold=True, size=11)
            r += 1
            for diag in diagnostics:
                ws_detail.cell(row=r, column=1, value=diag.get("symptom", "")).font = Font(bold=True, color="DC2626")
                r += 1
                for cause in diag.get("causes", []):
                    ws_detail.cell(row=r, column=1, value=f"  원인: {cause.get('cause', '')}").border = thin_border
                    ws_detail.cell(row=r, column=2, value=f"확인: {cause.get('check', '')}").border = thin_border
                    r += 1
                r += 1

        # 열 너비
        ws_detail.column_dimensions["A"].width = 35
        ws_detail.column_dimensions["B"].width = 25
        ws_detail.column_dimensions["C"].width = 20
        ws_detail.column_dimensions["D"].width = 15
        ws_detail.column_dimensions["E"].width = 40

    # ── 저장 ────────────────────────────────────────────────
    output_name = f"summary_{today}.xlsx"
    output_path = Path(result_dir) / output_name
    wb.save(str(output_path))
    return output_path


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 generate-summary.py <result_dir>")
        print("  예: python3 generate-summary.py results/2026-03-19")
        sys.exit(1)

    result_dir = sys.argv[1]
    if not os.path.isdir(result_dir):
        print(f"  ⚠  디렉토리 없음: {result_dir}")
        sys.exit(1)

    results = load_json_results(result_dir)
    if not results:
        print(f"  ⚠  JSON 결과 파일 없음: {result_dir}")
        sys.exit(1)

    print(f"  JSON 파일 {len(results)}개 로드")
    for r in results:
        scenario = r.get("scenario", "unknown")
        pass_text = "PASS" if r.get("pass") else "FAIL"
        print(f"    - {scenario}: {pass_text}")

    output_path = create_excel(results, result_dir)
    print(f"  → Excel 생성 완료: {output_path}")


if __name__ == "__main__":
    main()
